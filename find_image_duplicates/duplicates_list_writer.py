import os.path
from typing import List

from difPy import dif
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def get_path(raw_path: str) -> str:    
    return raw_path.replace('\\', '/')


def get_file_name(path: str) -> str:
    res = path.split('\\')
    return res[-1]


def search_duplicates(folder1: str, folder2: str=None) -> dict:
    search = dif(folder1, folder2)
    return search.result


def write_headers(sheet: Worksheet, headers: List[str], row: int=1) -> None:    
    for header_pos in range(len(headers)):
        cell = sheet.cell(row = row, column = header_pos+1)
        cell.value = headers[header_pos]


def fill_cell(sheet: Worksheet, row: int, col: int, value: str) -> None:
    cell = sheet.cell(row = row, column = col+1)
    cell.value = value


def fill_col(
        sheet: Worksheet, 
        headers: List[str], 
        row: int, 
        col: int, 
        values: List[str],
        processed: set,            
    ) -> int:              
            
    if headers[col] != headers[-1]:
        fill_cell(sheet, row, col, values[0])
        return row, processed
    
    for value in values:                    
        fill_cell(sheet, row, col, value)        
        row += 1
        processed.add(get_file_name(value))
    return row, processed


def write_body(sheet: Worksheet, headers: list, found_duplicates: dict, row: int=2) -> None:    
    processed = set()
    for _, val in found_duplicates:
        if val[headers[0]] in processed:            
            continue    
        
        for col in range(len(headers)):
            content = val[headers[col]]    
            values = content if isinstance(content, list) else [content] 
            row, processed = fill_col(sheet, headers, row, col, values, processed)            
  

def name_validation(file_name: str) -> str:
    if not os.path.isfile(file_name):
        return file_name
    
    split_name = file_name.split('.')
    new_name = split_name[0] + '_copy.' + split_name[1]
    return name_validation(new_name)


def write_resaults(
        search_resault: dict, 
        file_name: str='find_image_duplicates/duplicates.xlsx',
    ) -> None:

    wb = Workbook()
    sheet = wb.active    

    headers = ['filename', 'location', 'duplicates']
    write_headers(sheet, headers)    
    write_body(sheet, headers, search_resault.items())

    valid_name = name_validation(file_name)
    wb.save(valid_name)


def is_different(ph1: str, ph2: str, pos: int=0):    
    ph1, ph2 = ph1.split('\\'), ph2.split('\\')    

    while pos < len(min(ph1, ph2)):
        if ph1[pos] != ph2[pos]:            
            return True
        pos += 1

    return False
    

def path_name_validation(paths: List[str]) -> None:    
    if len(paths) == 1:
        return
    if len(paths) != 2:
        raise ValueError
    if len(set(paths)) == 1:
        raise ValueError
    if not is_different(*paths):
        raise ValueError


def path_validation(paths: List[str]) -> List[str]:
    try:
        path_name_validation(paths)
    except ValueError:
        print('Not allowed path combination')
        return

    for path in paths:
        if not os.path.isdir(path):
            print(f'Path is not exist: {path}')
            return

    return [get_path(path) for path in paths]
    

def main():
    paths = [
        r'C:\Users\alex-\OneDrive\Изображения\vadim\камера iphone\iCloud Photos',
        r'C:\Users\alex-\OneDrive\Изображения\vadim\камера iphone',
        
    ]
    
    validated_paths = path_validation(paths)
    if not validated_paths:
        return

    search_resault = search_duplicates(*validated_paths)        
    write_resaults(search_resault)


if __name__ == '__main__':
    main()
