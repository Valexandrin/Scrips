import os
from typing import List

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


def get_cells_by_poor_resolution(worksheet: Worksheet) -> List[Cell]:
    cells = []
    for row in worksheet.rows:
        path_cells = []
        for cell in row:
            if os.path.isfile(cell.value):
                path_cells.append(cell)

        if path_cells:
            best_res = min(path_cells, key=lambda cell: os.stat(cell.value).st_size)
            cells.append(best_res)

    return cells


def file_markup(cells: List[Cell]) -> None:
    for cell in cells:
        cell.fill = PatternFill("solid", fgColor="ffff00")


def main():
    fname = 'find_image_duplicates/duplicates_copy_copy.xlsx'
    wb = load_workbook(fname)
    ws = wb.active

    cells = get_cells_by_poor_resolution(ws)
    file_markup(cells)

    wb.save(fname)


if __name__ == '__main__':
    main()
