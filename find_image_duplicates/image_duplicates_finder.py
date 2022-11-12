from typing import List

from difPy import dif
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def get_path(raw_path: str) -> str:    
    return raw_path.replace('\\', '/')


def get_file_name(path: str) -> str:
    res = path.split('\\')
    return res[-1]


def search_duplicates(folder1: str, folder2: str=None) -> dict:
    search = dif(folder1, folder2)
    return search.result


def write_headers(sheet: Worksheet, headers: List[str], row: int=1) -> None:    
    for header_pos in range(len(headers)):
        cell = sheet.cell(row = row, column = header_pos+1)
        cell.value = headers[header_pos]


def fill_cell(sheet: Worksheet, row: int, col: int, value: str) -> None:
    cell = sheet.cell(row = row, column = col+1)
    cell.value = value


def fill_col(sheet: Worksheet, 
            headers: List[str], 
            row: int, 
            col: int, 
            values: List[str],
            processed: set,            
    ) -> int:              
            
    if headers[col] != headers[-1]:
        fill_cell(sheet, row, col, values[0])
        return row, processed
    
    for value in values:                    
        fill_cell(sheet, row, col, value)        
        row += 1
        processed.add(get_file_name(value))
    return row, processed


def write_body(sheet: Worksheet, headers: list, found_duplicates: dict, row: int=2) -> None:    
    processed = set()
    for _, val in found_duplicates:
        if val[headers[0]] in processed:            
            continue    
        
        for col in range(len(headers)):
            content = val[headers[col]]    
            values = content if isinstance(content, list) else [content] 
            row, processed = fill_col(sheet, headers, row, col, values, processed)            
  

def write_resaults(search_resault: dict, file_name: str) -> None:
    wb = Workbook()
    sheet = wb.active    

    headers = ['filename', 'location', 'duplicates']
    write_headers(sheet, headers)    
    write_body(sheet, headers, search_resault.items())
                
    wb.save(file_name)


def main():
    path = get_path(r'C:\Users\alex-\OneDrive\Изображения\vadim\камера iphone\iCloud Photos')
    search_resault = search_duplicates(path)

    resault_file_name = 'duplicates.xlsx'
    write_resaults(search_resault, resault_file_name)


if __name__ == '__main__':
    main()
